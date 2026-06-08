from __future__ import annotations

import csv
import math
import re
import unicodedata
from dataclasses import dataclass
from pathlib import Path

from PIL import Image, ImageDraw, ImageFont
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side


MONTH_LABEL = "IUNIE 2026"
CHAPTER_LABEL = "CREATIV"
INPUT_CSV = Path(
    r"G:\My Drive\BNI\BNI - Professional\Regiuni - CJ\Grupuri\CREATIV\Semafor\2026\06 IUN\Rank.csv"
)
OUTPUT_DIR = INPUT_CSV.parent

GREEN = "#a9dec4"
YELLOW = "#ffe67d"
RED = "#ef9aa7"
GRAY = "#c7c7c7"
WHITE = "#ffffff"
GRID = "#e2e2e2"
TEXT = "#222222"
TEAL = "#188aa0"
EC_GREEN = "#91d04e"
EC_YELLOW = "#fff500"
EC_RED = "#f18782"


def font(size: int, bold: bool = False) -> ImageFont.FreeTypeFont:
    candidates = [
        r"C:\Windows\Fonts\arialbd.ttf" if bold else r"C:\Windows\Fonts\arial.ttf",
        r"C:\Windows\Fonts\segoeuib.ttf" if bold else r"C:\Windows\Fonts\segoeui.ttf",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"
        if bold
        else "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
    ]
    for candidate in candidates:
        if Path(candidate).exists():
            return ImageFont.truetype(candidate, size=size)
    return ImageFont.load_default()


F = {
    "title": font(74, True),
    "head": font(20, True),
    "cell": font(20),
    "cell_small": font(18),
    "compact_head": font(19, True),
    "compact_cell": font(19),
    "compact_title": font(86, True),
    "ec": font(28),
    "ec_bold": font(26, True),
}


@dataclass
class Member:
    rank: int
    region: str
    chapter: str
    name: str
    total: int
    attendance_rate: float
    attendance_score: int
    ceu_rate: float
    ceu_score: int
    one_rate: float
    one_score: int
    referrals_rate: float
    referrals_score: int
    tyfcb: int
    tyfcb_score: int
    visitors: int
    visitor_score: int
    sponsors: int
    sponsor_score: int


def fnum(value: str) -> float:
    return float((value or "0").replace(",", ""))


def read_members() -> list[Member]:
    with INPUT_CSV.open("r", encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.DictReader(handle))
    members = []
    for row in rows:
        members.append(
            Member(
                rank=int(fnum(row["Rank"])),
                region=row["Region"].strip(),
                chapter=row["Chapter"].strip(),
                name=re.sub(r"\s+", " ", row["Name"].strip()),
                total=int(fnum(row["Total Score"])),
                attendance_rate=fnum(row["Attendance Rate"]),
                attendance_score=int(fnum(row["Attendance Score"])),
                ceu_rate=fnum(row["CEU Rate"]),
                ceu_score=int(fnum(row["CEU Score"])),
                one_rate=fnum(row["121 Rate"]),
                one_score=int(fnum(row["121 Score"])),
                referrals_rate=fnum(row["Referrals Given Rate"]),
                referrals_score=int(fnum(row["Referrals Given Score"])),
                tyfcb=int(fnum(row["TYFCB"])),
                tyfcb_score=int(fnum(row["TYFCB Score"])),
                visitors=int(fnum(row["Visitors"])),
                visitor_score=int(fnum(row["Visitor Score"])),
                sponsors=int(fnum(row["Sponsors"])),
                sponsor_score=int(fnum(row["Sponsor Score"])),
            )
        )
    return sorted(members, key=lambda m: m.rank)


def normalize_name(value: str) -> str:
    value = value.replace("ă", "a").replace("â", "a").replace("î", "i").replace("ș", "s").replace("ş", "s")
    value = value.replace("ț", "t").replace("ţ", "t").replace("-", " ")
    value = unicodedata.normalize("NFKD", value).encode("ascii", "ignore").decode("ascii")
    return re.sub(r"\s+", " ", value).strip().lower()


def wrap_text(draw: ImageDraw.ImageDraw, text: str, max_width: int, face: ImageFont.FreeTypeFont) -> list[str]:
    words = str(text).split()
    if not words:
        return [""]
    lines: list[str] = []
    current = words[0]
    for word in words[1:]:
        candidate = f"{current} {word}"
        if draw.textbbox((0, 0), candidate, font=face)[2] <= max_width:
            current = candidate
        else:
            lines.append(current)
            current = word
    lines.append(current)
    return lines


def draw_wrapped(
    draw: ImageDraw.ImageDraw,
    xy: tuple[int, int],
    text: str,
    width: int,
    face: ImageFont.FreeTypeFont,
    fill: str = TEXT,
    line_gap: int = 4,
    bold_first: bool = False,
) -> None:
    x, y = xy
    lines = wrap_text(draw, text, width, face)
    for line in lines:
        draw.text((x, y), line, font=face, fill=fill)
        y += face.size + line_gap


def color_total(score: int) -> str:
    if score >= 70:
        return GREEN
    if score >= 50:
        return YELLOW
    if score >= 30:
        return RED
    return GRAY


def color_score(score: int, max_score: int) -> str:
    if score <= 0:
        return GRAY if max_score in (20, 25, 5) and score == 0 else RED
    ratio = score / max_score
    if ratio >= 0.8:
        return GREEN
    if ratio >= 0.4:
        return YELLOW
    return RED


def draw_sort_icon(draw: ImageDraw.ImageDraw, x: int, y: int, scale: int = 5) -> None:
    c = "#d7d7d7"
    draw.polygon([(x, y), (x + scale, y - scale), (x + 2 * scale, y)], fill=c)
    draw.polygon([(x, y + 8), (x + scale, y + 8 + scale), (x + 2 * scale, y + 8)], fill=c)


def cell(
    draw: ImageDraw.ImageDraw,
    x: int,
    y: int,
    w: int,
    h: int,
    text: str,
    fill: str = WHITE,
    face: ImageFont.FreeTypeFont | None = None,
    align: str = "center",
    text_fill: str = TEXT,
    border: bool = True,
) -> None:
    face = face or F["cell"]
    draw.rectangle((x, y, x + w, y + h), fill=fill)
    if border:
        draw.line((x, y + h, x + w, y + h), fill=GRID, width=1)
        draw.line((x + w, y, x + w, y + h), fill=GRID, width=1)
    lines = wrap_text(draw, text, max(1, w - 14), face)
    total_h = len(lines) * face.size + (len(lines) - 1) * 3
    ty = y + (h - total_h) // 2 - 1
    for line in lines:
        bbox = draw.textbbox((0, 0), line, font=face)
        tw = bbox[2] - bbox[0]
        if align == "left":
            tx = x + 10
        elif align == "right":
            tx = x + w - tw - 10
        else:
            tx = x + (w - tw) // 2
        draw.text((tx, ty), line, font=face, fill=text_fill)
        ty += face.size + 3


def format_value(member: Member, key: str) -> str:
    if key == "rank":
        return str(member.rank)
    if key == "region":
        return member.region
    if key == "chapter":
        return member.chapter.replace(" ", "\n")
    if key == "name":
        return member.name
    if key == "total":
        return str(member.total)
    if key == "attendance_rate":
        return f"{member.attendance_rate * 100:.1f}%"
    if key == "attendance_score":
        return str(member.attendance_score)
    if key == "ceu_rate":
        return f"{member.ceu_rate:.1f}"
    if key == "ceu_score":
        return str(member.ceu_score)
    if key == "one_rate":
        return f"{member.one_rate:.2f}"
    if key == "one_score":
        return str(member.one_score)
    if key == "referrals_rate":
        return f"{member.referrals_rate:.2f}"
    if key == "referrals_score":
        return str(member.referrals_score)
    if key == "tyfcb":
        return f"{member.tyfcb:,}"
    if key == "tyfcb_score":
        return str(member.tyfcb_score)
    if key == "visitors":
        return str(member.visitors)
    if key == "visitor_score":
        return str(member.visitor_score)
    if key == "sponsors":
        return str(member.sponsors)
    if key == "sponsor_score":
        return str(member.sponsor_score)
    raise KeyError(key)


def metric_color(member: Member, key: str) -> str:
    if key in {"rank", "region", "chapter", "name", "tyfcb"}:
        return WHITE
    if key == "total":
        return color_total(member.total)
    if key in {"attendance_rate", "attendance_score"}:
        return color_score(member.attendance_score, 10)
    if key in {"ceu_rate", "ceu_score"}:
        return color_score(member.ceu_score, 10)
    if key in {"one_rate", "one_score"}:
        return color_score(member.one_score, 20)
    if key in {"referrals_rate", "referrals_score"}:
        return color_score(member.referrals_score, 25)
    if key == "tyfcb_score":
        return color_score(member.tyfcb_score, 5)
    if key in {"visitors", "visitor_score"}:
        return color_score(member.visitor_score, 25)
    if key in {"sponsors", "sponsor_score"}:
        return GREEN if member.sponsor_score >= 5 else RED
    return WHITE


def render_detailed(members: list[Member]) -> Path:
    cols = [
        ("rank", "Rank", 70),
        ("region", "Region", 85),
        ("chapter", "Chapter", 115),
        ("name", "Name", 195),
        ("total", "Total\nScore", 130),
        ("attendance_rate", "Attendance\nRate", 175),
        ("attendance_score", "Attendance\nScore", 185),
        ("ceu_rate", "CEU\nRate", 130),
        ("ceu_score", "CEU\nScore", 135),
        ("one_rate", "121\nRate", 135),
        ("one_score", "121\nScore", 135),
        ("referrals_rate", "Referrals\nGiven Rate", 230),
        ("referrals_score", "Referrals\nGiven Score", 230),
        ("tyfcb", "TYFCB", 125),
        ("tyfcb_score", "TYFCB\nScore", 125),
        ("visitors", "Visitors", 120),
        ("visitor_score", "Visitor\nScore", 155),
        ("sponsors", "Sponsors", 120),
        ("sponsor_score", "Sponsor\nScore", 158),
    ]
    width = sum(c[2] for c in cols)
    row_h, head_h = 92, 174
    height = head_h + row_h * len(members)
    img = Image.new("RGB", (width, height), WHITE)
    draw = ImageDraw.Draw(img)
    origin_y = 0
    x = 0
    for key, label, w in cols:
        cell(draw, x, origin_y, w, head_h, label, WHITE, F["head"].font_variant(size=23), "center")
        draw_sort_icon(draw, x + w - 22, origin_y + 100, 5)
        x += w
    y = head_h
    for member in members:
        x = 0
        for key, _label, w in cols:
            face = F["cell_small"].font_variant(size=21) if key in {"chapter", "name"} else F["cell"].font_variant(size=23)
            text_fill = TEAL if key == "name" else TEXT
            align = "left" if key == "name" else "center"
            cell(
                draw,
                x,
                y,
                w,
                row_h,
                format_value(member, key),
                metric_color(member, key),
                face,
                align,
                text_fill,
            )
            x += w
        y += row_h
    for line_y in range(head_h, height + 1, row_h):
        draw.line((0, line_y, width, line_y), fill="#cfcfcf", width=2)
    out = OUTPUT_DIR / f"SEMAFOR {CHAPTER_LABEL} - {MONTH_LABEL}.jpg"
    img.save(out, quality=94, subsampling=1)
    return out


def render_compact(members: list[Member]) -> Path:
    img = Image.new("RGB", (3266, 2412), WHITE)
    draw = ImageDraw.Draw(img)
    title = f"SEMAFOR {MONTH_LABEL}"
    title_font = F["compact_title"].font_variant(size=110)
    tw = draw.textbbox((0, 0), title, font=title_font)[2]
    draw.text(((3266 - tw) // 2, 115), title, font=title_font, fill="#000000")
    cols = [("rank", "Rank", 88), ("region", "Region", 108), ("chapter", "Chapter", 150), ("name", "Name", 182), ("total", "Total\nScore", 122)]
    block_w = sum(c[2] for c in cols)
    x_positions = [180, 955, 1730, 2480]
    y_positions = [390, 560, 690, 840]
    groups = [
        [m for m in members if m.total >= 70],
        [m for m in members if 50 <= m.total < 70],
        [m for m in members if 30 <= m.total < 50],
        [m for m in members if m.total < 30],
    ]
    for block_idx, group in enumerate(groups):
        x0 = x_positions[block_idx]
        y0 = y_positions[block_idx]
        head_h = 68
        row_h = 90 if block_idx == 3 else 92
        x = x0
        for key, label, w in cols:
            cell(draw, x, y0, w, head_h, label, WHITE, F["compact_head"], "center")
            draw_sort_icon(draw, x + w - 20, y0 + 42, 4)
            x += w
        y = y0 + head_h
        for member in group:
            x = x0
            for key, _label, w in cols:
                face = F["compact_cell"]
                fill = color_total(member.total) if key == "total" else WHITE
                text_fill = TEAL if key == "name" else TEXT
                align = "left" if key == "name" else "center"
                cell(draw, x, y, w, row_h, format_value(member, key), fill, face, align, text_fill)
                x += w
            y += row_h
    out = OUTPUT_DIR / f"BNI {CHAPTER_LABEL} - Semafor - {MONTH_LABEL} (Membri).jpg"
    img.save(out, quality=94, subsampling=1)
    return out


EC_ROWS = [
    ("Președinte", "Corina Amalia Vaida"),
    ("Vicepreședinte", "Bianca Sharma"),
    ("Secretar / Trezorier", "Alexandru Ardelean"),
    ("Comitetul de Întâmpinare", "Andreea Badiu"),
    ("Comitetul de Întâmpinare", "Monica Teodora Parcalab"),
    ("Comitetul de Întâmpinare", "Radu Carzon"),
    ("Comitetul de Întâmpinare", "Tunde Maksay"),
    ("Coordonator educațional", "Cristina Bara"),
    ("Coordonator educațional", "Cristian Muresan"),
    ("Coordonator evenimente", "Anca Serban"),
    ("Coordonator evenimente", "Mariana Radulescu"),
    ("Coordonator mentori", "Anca Roșu"),
    ("Coordonator mentori", "Daniela Șerban"),
    ("Comitetul de Membri", "Romina Alina Marc"),
    ("Comitetul de Membri", "Adelina Muntean"),
    ("Comitetul de Membri", "Claudiu Pop"),
    ("Comitetul de Membri", "Ioana Rogozan"),
    ("Membership Committee - Community Builder", ""),
    ("Membership Committee - Quality Assurance", ""),
    ("Membership Committee - Member Engagement", ""),
    ("Membership Committee - Member Relations", ""),
    ("Coordonator Dezvoltare", "Ioana Filimon"),
    ("Coordonator Dezvoltare", "Cleopatra Catana"),
    ("Coordonator PowerTeam", "Ana Maria Mociofan"),
]


def ec_bar_color(member: Member | None) -> str:
    if member is None:
        return WHITE
    if member.total >= 70:
        return EC_GREEN
    if member.total >= 50:
        return EC_YELLOW
    if member.total >= 30:
        return EC_RED
    return GRAY


def ec_comment(member: Member | None) -> tuple[str, str]:
    return WHITE, ""


def render_ec(members: list[Member]) -> Path:
    by_name = {normalize_name(m.name): m for m in members}
    aliases = {
        normalize_name("Ana Maria Mociofan"): normalize_name("Ana-Maria Mociofan"),
        normalize_name("Ioana Filimon"): normalize_name("Ioana Filimon"),
        normalize_name("Mariana Radulescu"): normalize_name("Mariana Radulescu"),
        normalize_name("Andreea Badiu"): normalize_name("Andreea Badiu"),
    }
    col_w = [920, 380, 250, 988]
    img = Image.new("RGB", (sum(col_w[:3]), 1070), WHITE)
    draw = ImageDraw.Draw(img)
    row_h = 45
    y = 0
    for idx, (role, name) in enumerate(EC_ROWS):
        key = normalize_name(name)
        member = by_name.get(key) or by_name.get(aliases.get(key, ""))
        bar_fill = ec_bar_color(member)
        comment_fill, comment = ec_comment(member)
        x = 0
        bg = "#f5f7f7" if idx % 2 == 0 else WHITE
        cell(draw, x, y, col_w[0], row_h, role, bg, F["ec_bold"], "left")
        x += col_w[0]
        cell(draw, x, y, col_w[1], row_h, name, bg, F["ec"], "left")
        x += col_w[1]
        cell(draw, x, y, col_w[2], row_h, "", bar_fill, F["ec"], "center")
        y += row_h
    out = OUTPUT_DIR / f"SEMAFOR {CHAPTER_LABEL} - {MONTH_LABEL} (Echipa Conducere).jpg"
    img.save(out, quality=94, subsampling=1)
    return out


def hex_no_hash(value: str) -> str:
    return value.replace("#", "").upper()


def export_ec_xlsx(members: list[Member]) -> Path:
    by_name = {normalize_name(m.name): m for m in members}
    aliases = {
        normalize_name("Ana Maria Mociofan"): normalize_name("Ana-Maria Mociofan"),
        normalize_name("Ioana Filimon"): normalize_name("Ioana Filimon"),
        normalize_name("Mariana Radulescu"): normalize_name("Mariana Radulescu"),
        normalize_name("Andreea Badiu"): normalize_name("Andreea Badiu"),
    }
    wb = Workbook()
    ws = wb.active
    ws.title = "EC"
    widths = {"A": 58, "B": 27, "C": 15, "D": 75}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row_idx, (role, name) in enumerate(EC_ROWS, 1):
        key = normalize_name(name)
        member = by_name.get(key) or by_name.get(aliases.get(key, ""))
        bar_fill = ec_bar_color(member)
        comment_fill, comment = ec_comment(member)
        bg = "#F5F7F7" if row_idx % 2 == 1 else WHITE
        values = [role, name, "", comment]
        fills = [bg, bg, bar_fill, comment_fill]
        for col_idx, (value, fill) in enumerate(zip(values, fills), 1):
            c = ws.cell(row=row_idx, column=col_idx, value=value)
            c.fill = PatternFill("solid", fgColor=hex_no_hash(fill))
            c.border = border
            c.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)
            c.font = Font(name="Arial", size=16, bold=col_idx == 1)
        ws.row_dimensions[row_idx].height = 26
    out = OUTPUT_DIR / "Semafor EC.xlsx"
    wb.save(out)
    return out


def main() -> None:
    members = read_members()
    outputs = [render_detailed(members), render_compact(members), render_ec(members), export_ec_xlsx(members)]
    print("\n".join(str(p) for p in outputs))


if __name__ == "__main__":
    main()
